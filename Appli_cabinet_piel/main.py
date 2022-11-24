from openpyxl import *
from tkinter import *
#from list_partenaire import *
#from liste_Dossier import *

wb = load_workbook("data.xlsx")

sheet = wb.active


def excel():
    sheet.column_dimensions['A'].width = 20  # larger column A
    sheet.column_dimensions['B'].width = 20  # larger column B
    sheet.column_dimensions['C'].width = 20  # larger column C
    sheet.column_dimensions['D'].width = 20  # larger column D
    sheet.column_dimensions['E'].width = 20  # larger column E
    sheet.column_dimensions['F'].width = 20  # larger column F
    sheet.column_dimensions['G'].width = 20  # larger column G
    sheet.column_dimensions['H'].width = 20  # larger column H
    sheet.column_dimensions['I'].width = 20  # larger column I
    sheet.column_dimensions['J'].width = 20  # larger column J
    sheet.column_dimensions['K'].width = 20  # larger column K
    sheet.column_dimensions['L'].width = 20  # larger column L

    sheet.cell(row=1, column=1).value = "Nom"  # column A
    sheet.cell(row=1, column=2).value = "Prénom"  # column B
    sheet.cell(row=1, column=3).value = "âge"  # column C
    sheet.cell(row=1, column=4).value = "Montant Trésorerie"  # column D
    sheet.cell(row=1, column=5).value = "Nombres de lignes de prêt"  # column E
    sheet.cell(row=1, column=6).value = "Montant du RAC"  # column F
    sheet.cell(row=1, column=7).value = "Durée du RAC"  # column G
    sheet.cell(row=1, column=8).value = "Catégorie du dossier"  # column H
    sheet.cell(row=1, column=9).value = "Pourcentage des honoraires"  # column I
    sheet.cell(row=1, column=10).value = "Montant des honoraires"  # column J
    sheet.cell(row=1, column=11).value = "Apporteur"  # column K
    sheet.cell(row=1, column=12).value = "Financeable"  # column L


def focus1():
    nom_field.focus_set()  # Focus on the last_name input field


def focus2():
    prenom_field.focus_set()  # Focus on the first_name input field


def focus3():
    age_field.focus_set()  # Focus on the age input field


def focus4():
    tresorerie_field.focus_set()  # Focus on the cashflow input field


def focus5():
    pret_field.focus_set()  # Focus on the loan lines input field


def focus6():
    montant_rac_field.focus_set()  # Focus on the values of loan input field


def focus7():
    duree_rac_field.focus_set()  # Focus on the duration of the loan input field


def focus8():
    cat_dossier_field.focus_set()  # Focus on the  type of case input field


def focus9():
    pourcentage_honoraires_field.focus_set()  # Focus on the percentage of fees input field


def focus10():
    montant_honoraires_field.focus_set()  # Focus on the amount of the fees input field


def focus11():
    apporteur_field.focus_set()  # Focus on the origin of the field  input field


def focus12():
    financeable_field.focus_set()  # Focus on the last_name input field


def clear():
    nom_field.delete(0, END)
    prenom_field.delete(0, END)
    age_field.delete(0, END)
    tresorerie_field.delete(0, END)
    pret_field.delete(0, END)
    montant_rac_field.delete(0, END)
    duree_rac_field.delete(0, END)
    cat_dossier_field.delete(0, END)
    pourcentage_honoraires_field.delete(0, END)
    montant_honoraires_field.delete(0, END)
    apporteur_field.delete(0, END)


def insert():
    if (nom_field.get() == "" and
            prenom_field.get() == "" and
            age_field.get() == "" and
            tresorerie_field.get() == "" and
            pret_field.get() == "" and
            montant_rac_field.get() == "" and
            duree_rac_field.get() == "" and
            cat_dossier_field.get() == "" and
            pourcentage_honoraires_field.get() == "" and
            montant_honoraires_field.get() == "" and
            apporteur_field.get() == ""):

        print("entrez des données")
    else:

        current_row = sheet.max_row

        sheet.cell(row=current_row + 1, column=1).value = nom_field.get()
        sheet.cell(row=current_row + 1, column=2).value = prenom_field.get()
        sheet.cell(row=current_row + 1, column=3).value = age_field.get()
        sheet.cell(row=current_row + 1, column=4).value = tresorerie_field.get()
        sheet.cell(row=current_row + 1, column=5).value = pret_field.get()
        sheet.cell(row=current_row + 1, column=6).value = montant_rac_field.get()
        sheet.cell(row=current_row + 1, column=7).value = duree_rac_field.get()
        sheet.cell(row=current_row + 1, column=8).value = cat_dossier_field.get()
        sheet.cell(row=current_row + 1, column=9).value = pourcentage_honoraires_field.get()
        sheet.cell(row=current_row + 1, column=10).value = montant_honoraires_field.get()
        sheet.cell(row=current_row + 1, column=11).value = apporteur_field.get()

        wb.save("data.xlsx")

        nom_field.focus_set()

        clear()


if __name__ == "__main__":
    root = Tk()

    root.configure(background='#f2f2f2')

    root.title("Cabinet Piel Courtage")

    root.geometry("900x700")

    excel()

heading = Label(text="Suivi CA", bg='#f2f2f2', fg='black', font=('Arial', 30, 'bold'))
nom = Label(text="Nom", bg='#f2f2f2', fg='black', font=('Arial', 20, 'bold'))
prenom = Label(text="Prénom", bg='#f2f2f2', fg='black', font=('Arial', 20, 'bold'))
age = Label(text="Âge", bg='#f2f2f2', fg='black', font=('Arial', 20, 'bold'))
tresorerie = Label(text="Trésorerie", bg='#f2f2f2', fg='black', font=('Arial', 20, 'bold'))
pret = Label(text="Nombre de lignes de prêt", bg='#f2f2f2', fg='black', font=('Arial', 20, 'bold'))
montant_rac = Label(text="Montant du RAC", bg='#f2f2f2', fg='black', font=('Arial', 20, 'bold'))
duree_rac = Label(text="Durée du RAC (en mois)", bg='#f2f2f2', fg='black', font=('Arial', 20, 'bold'))
dossier = Label(text="Type de dossier", bg='#f2f2f2', fg='black', font=('Arial', 20, 'bold'))
pourcentage_dossier = Label(text="Pourcentage des honoraires", bg='#f2f2f2', fg='black', font=('Arial', 20, 'bold'))
montant_honoraires = Label(text="Montant des honoraires", bg='#f2f2f2', fg='black', font=('Arial', 20, 'bold'))
apporteur = Label(text="Apporteur", bg='#f2f2f2', fg='black', font=('Arial', 20, 'bold'))
financeable = Label(text="Financeable", bg='#f2f2f2', fg='black', font=('Arial', 20, 'bold'))

heading.grid(row=0, column=1)
nom.grid(row=1, column=0)
prenom.grid(row=2, column=0)
age.grid(row=3, column=0)
tresorerie.grid(row=4, column=0)
pret.grid(row=5, column=0)
montant_rac.grid(row=6, column=0)
duree_rac.grid(row=7, column=0)
dossier.grid(row=8, column=0)
pourcentage_dossier.grid(row=9, column=0)
montant_honoraires.grid(row=10, column=0)
apporteur.grid(row=11, column=0)
financeable.grid(row=12, column=0)

nom_field = Entry(root)
prenom_field = Entry(root)
age_field = Entry(root)
tresorerie_field = Entry(root)
pret_field = Entry(root)
montant_rac_field = Entry(root)
duree_rac_field = Entry(root)
cat_dossier_field = Entry(root)
pourcentage_honoraires_field = Entry(root)
montant_honoraires_field = Entry(root)
apporteur_field = Entry(root)
financeable_field = Entry(root)

nom_field.bind("<Return>", focus1)
prenom_field.bind("<Return>", focus2)
age_field.bind("<Return>", focus3)
tresorerie_field.bind("<Return>", focus4)
pret_field.bind("<Return>", focus5)
montant_rac_field.bind("<Return>", focus6)
duree_rac_field.bind("<Return>", focus7)
cat_dossier_field.bind("<Return>", focus8)
pourcentage_honoraires_field.bind("<Return>", focus9)
montant_honoraires_field.bind("<Return>", focus10)
apporteur_field.bind("<Return>", focus11)
financeable_field.bind("<Return>", focus12)

nom_field.grid(row=1, column=1, ipadx="100")
prenom_field.grid(row=2, column=1, ipadx="100")
age_field.grid(row=3, column=1, ipadx="100")
tresorerie_field.grid(row=4, column=1, ipadx="100")
pret_field.grid(row=5, column=1, ipadx="100")
montant_rac_field.grid(row=6, column=1, ipadx="100")
duree_rac_field.grid(row=7, column=1, ipadx="100")
cat_dossier_field.grid(row=8, column=1, ipadx="100")
pourcentage_honoraires_field.grid(row=9, column=1, ipadx="100")
montant_honoraires_field.grid(row=10, column=1, ipadx="100")
apporteur_field.grid(row=11, column=1, ipadx="100")
financeable_field.grid(row=12, column=1, ipadx="100")

excel()

submit = Button(root, text="Envoi", fg="Black", bg="White", command=insert)
submit.grid(row=13, column=1)

root.mainloop()
